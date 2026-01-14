from fastapi import FastAPI, HTTPException, Body
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
import sqlite3
from pathlib import Path
import shutil
import string
from openai import OpenAI
from fastapi import UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse
import io
import os
import base64
import random
import uuid
import csv
import zipfile
# optional libs for Excel/PDF parsing
try:
    import openpyxl
except Exception:
    openpyxl = None
try:
    import PyPDF2
except Exception:
    PyPDF2 = None
from email.message import EmailMessage
from email import policy
from email.utils import format_datetime
import re
import datetime

EXAMPLES_DIR = Path(__file__).parent / "examples" / "pdfs"
STORAGE_DIR = Path(__file__).parent / "storage"
STORAGE_DOCS_DIR = STORAGE_DIR / "documents"
STORAGE_EMAILS_DIR = STORAGE_DIR / "emails"
# central inbox that all mails are sent to (for sorting)
INBOX_ADDRESS = "inbox@spitex.local"

# optional PDF generator
try:
    from reportlab.pdfgen import canvas
except Exception:
    canvas = None

# OpenAI client setup
client = OpenAI(
    base_url='http://127.0.0.1:1234/v1',
    api_key='test-api-key'
)

app = FastAPI()

# Pydantic models
class MessageSchema(BaseModel):
    client_id: str
    sender_id: str
    recipient_id: str
    content: str | None = None
    message_type: str = "message"
    is_important: bool = False
    due_date: str | None = None
    status: str = "pending"
    file_path: str | None = None
    file_name: str | None = None


# CORS settings
origins = [
    'http://localhost:5173',
    "http://localhost:8080"
]

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# SQLite DB path (renamed to something meaningful)
DB_PATH = Path(__file__).parent / "spitex.db"


def _generate_code(length: int = 5) -> str:
    """Generate a numeric-only code of given length (digits only)."""
    alphabet = string.digits
    return ''.join(random.choice(alphabet) for _ in range(length))


def _resolve_patient(identifier: str):
    """Resolve either numeric id or patient code to a (id, code, name) tuple.
    Raises HTTPException(404) if not found.
    """
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        # If identifier is digits-only, it could be either a numeric patient id or a numeric code.
        # Try to resolve by code first (codes are fixed-length numeric strings), then fall back to id.
        if identifier.isdigit():
            cur.execute("SELECT id, code, name FROM patients WHERE code = ?", (identifier,))
            row = cur.fetchone()
            if row:
                return (row["id"], row["code"], row["name"])
            # fall back to numeric id
            cur.execute("SELECT id, code, name FROM patients WHERE id = ?", (int(identifier),))
            row = cur.fetchone()
            if row:
                return (row["id"], row["code"], row["name"])
        # otherwise try code (case-insensitive for legacy alphanumeric support)
        cur.execute("SELECT id, code, name FROM patients WHERE UPPER(code) = UPPER(?)", (identifier,))
        row = cur.fetchone()
        if row:
            return (row["id"], row["code"], row["name"])
        # not found
        raise HTTPException(status_code=404, detail=f"Patient not found: {identifier}")
    finally:
        conn.close()


def _ai_parse_csv_for_medlist(csv_text: str) -> str | None:
    """Call the AI to extract medication rows from a CSV-like text and return a CSV string.
    Returns None on failure or if AI does not find med rows.
    """
    if not csv_text or not csv_text.strip():
        return None
    try:
        prompt = (
            "You receive the contents of a CSV file.\n"
            "Identify rows that represent medication entries (drug name, dosage, frequency, quantity, notes).\n"
            "Return ONLY a valid CSV text (no commentary) that uses the SAME header row as the input if a header is present.\n"
            "If no header is present, create a single-column CSV with header 'line' and include the detected medication lines.\n"
            "If you cannot find medication rows, return an empty CSV with only the header row.\n\n"
            "INPUT CSV:\n" + csv_text[:4000]
        )
        response = client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[
                {"role": "system", "content": "You are a precise CSV extraction assistant. Output only CSV content."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.0,
            max_tokens=1500,
        )
        out = response.choices[0].message.content
        # strip code fences if present
        m = re.search(r"```(?:csv)?\n([\s\S]*?)```", out, re.I)
        csv_out = m.group(1).strip() if m else out.strip()
        lines = [l for l in csv_out.splitlines() if l.strip()]
        if not lines:
            return None
        try:
            rows = list(csv.reader(lines))
            out_buf = io.StringIO()
            writer = csv.writer(out_buf)
            for row in rows:
                writer.writerow(row)
            return out_buf.getvalue()
        except Exception:
            return csv_out
    except Exception:
        return None


def _extract_text_from_pdf_bytes(b: bytes) -> str | None:
    """Extract text from PDF bytes using PyPDF2 if available."""
    if not PyPDF2:
        return None
    try:
        reader = PyPDF2.PdfReader(io.BytesIO(b))
        pages = [p.extract_text() or "" for p in reader.pages]
        return "\n".join(pages)
    except Exception:
        return None


def _get_plain_body_from_message(msg) -> str:
    """Extract a best-effort plain-text body from an email.message.Message."""
    body_parts = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = part.get("Content-Disposition", "") or ""
            if ctype.startswith("text/") and 'attachment' not in disp:
                try:
                    body_parts.append(part.get_content())
                except Exception:
                    try:
                        body_parts.append(part.get_payload(decode=True).decode('utf-8', errors='replace'))
                    except Exception:
                        continue
    else:
        try:
            body_parts.append(msg.get_content())
        except Exception:
            try:
                body_parts.append(msg.get_payload(decode=True).decode('utf-8', errors='replace'))
            except Exception:
                pass
    text = "\n\n".join([p for p in body_parts if p])
    if not text:
        # try HTML fallback
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    try:
                        html = part.get_content()
                    except Exception:
                        html = part.get_payload(decode=True).decode('utf-8', errors='replace')
                    text = re.sub('<[^<]+?>', '', html)
                    break
    return text or ""


def _ai_extract_prescriptions(text: str, headers: list[str], patient_info: dict | None = None, doctor_info: dict | None = None, required_fields: list[str] | None = None) -> str | None:
    """Call AI to extract prescription rows from a free-text or CSV and return CSV text.
    The AI will be instructed to use the provided headers and to use patient_info and doctor_info
    to populate patient/doctor-related fields when possible. `required_fields` ensures the AI
    includes/returns these columns.
    Returns None on failure or when no prescriptions are found.
    """
    if not text or not text.strip():
        return None
    headers_line = ','.join(headers)
    patient_note = ''
    if patient_info:
        parts = []
        for k in ('id', 'code', 'name', 'address', 'age', 'dob'):
            if patient_info.get(k):
                parts.append(f"{k}: {patient_info.get(k)}")
        if parts:
            patient_note = 'Patient info: ' + '; '.join(parts) + '.'
    doctor_note = ''
    if doctor_info:
        parts = []
        for k in ('name', 'address', 'specialty'):
            if doctor_info.get(k):
                parts.append(f"{k}: {doctor_info.get(k)}")
        if parts:
            doctor_note = 'Doctor info: ' + '; '.join(parts) + '.'
    req_note = ''
    if required_fields:
        req_note = 'Required fields (must appear as columns): ' + ', '.join(required_fields) + '.'

    # build a simple synonyms map to help the AI map common labels to the template headers
    # common synonyms for German/English variations
    synonym_map = {
        'code': ['code', 'patient code', 'id', 'patient id'],
        'kunde': ['kunde', 'patient', 'patient name', 'patientenangaben', 'patient name'],
        'name': ['name', 'patient name', 'vorname', 'nachname'],
        'adresse': ['adresse', 'address'],
        'alter': ['alter', 'age'],
        'geburtsdatum': ['geburtsdatum', 'geb', 'dob', 'birthdate', 'geburtstag'],
        'medikament': ['medikament', 'drug', 'medication', 'rx', 'medicine', 'präparat'],
        'anzahl': ['anzahl', 'qty', 'quantity', 'menge'],
        'täglich': ['täglich', 'taeglich', 'daily', 'täglich?'],
        'morgens': ['morgens', 'morning'],
        'mittags': ['mittags', 'noon', 'midday'],
        'abends': ['abends', 'evening'],
        'behandelnder arzt': ['behandelnder arzt', 'arzt', 'doctor', 'prescriber'],
        'arzt spezialität': ['spezialität', 'specialty', 'speciality', 'arzt spezialität']
    }

    # include mapping hints in the prompt to bias AI output towards the template's (German) headers
    mapping_lines = []
    for std, sy in synonym_map.items():
        mapping_lines.append(f"{std}: {', '.join(sy)}")
    mapping_hint = "; ".join(mapping_lines)

    prompt = (
        f"You receive an arbitrary text (email body, OCR of a PDF or a CSV).\n"
        f"Extract all medication prescription entries you can find and return a VALID CSV using exactly these headers (in this order): {headers_line}.\n"
        "Use the German header names exactly as provided if possible. If you see common synonyms for a column, map them to the exact header name.\n"
        "For boolean fields like 'Täglich', use 'Ja' or 'Nein'. For frequency columns like 'Morgens', 'Mittags', 'Abends' put integers (0/1) or counts where appropriate.\n"
        "If you cannot find a value for a column, leave it empty. Ensure the CSV matches the requested headers exactly and includes the required fields if possible.\n"
        "Return ONLY the CSV text (no commentary). If no prescriptions are found, return a CSV with only the header row.\n\n"
        f"Synonym map (helps you map columns to headers): {mapping_hint}\n\n"
        f"{patient_note} {doctor_note} {req_note}\n"
        "INPUT:\n" + text[:12000]
    )

    try:
        response = client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[
                {"role": "system", "content": "You are a precise data extraction assistant. Output ONLY CSV with the requested headers."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.0,
            max_tokens=3200,
        )
        out = response.choices[0].message.content
        m = re.search(r"```(?:csv)?\n([\s\S]*?)```", out, re.I)
        csv_text = m.group(1).strip() if m else out.strip()

        # Normalize and try to remap AI-produced headers to the exact template headers
        def normalize_ai_csv(csv_text: str, template_headers: list[str]) -> str:
            lines = [l for l in csv_text.splitlines() if l.strip()]
            if not lines:
                return ''
            try:
                rows = list(csv.reader(lines))
            except Exception:
                # fallback: return raw text
                return csv_text
            ai_headers = [h.strip() for h in rows[0]] if rows else []
            # if headers already match exactly, return original
            if ai_headers and [h.strip() for h in ai_headers] == [h.strip() for h in template_headers]:
                out_buf = io.StringIO()
                writer = csv.writer(out_buf)
                for r in rows:
                    writer.writerow([str(c) for c in r])
                return out_buf.getvalue()

            # build reverse synonym lookup for matching
            rev_map = {}
            for std, syns in synonym_map.items():
                for s in syns:
                    rev_map[s.lower()] = std

            ai_lower = [h.lower() for h in ai_headers]
            mapping = {}
            for dest_idx, desired in enumerate(template_headers):
                desired_l = desired.strip().lower()
                # direct find
                found = None
                if desired_l in ai_lower:
                    found = ai_lower.index(desired_l)
                else:
                    # check synonyms
                    for j, ah in enumerate(ai_lower):
                        # exact synonym
                        if ah in rev_map and rev_map[ah] == desired_l:
                            found = j
                            break
                        # fuzzy contains
                        if desired_l in ah or ah in desired_l:
                            found = j
                            break
                    # also check if desired label contains a known synonym word
                    if not found:
                        for syn, std in rev_map.items():
                            if std == desired_l and syn in ' '.join(ai_lower):
                                # find ai header containing the syn
                                for j, ah in enumerate(ai_lower):
                                    if syn in ah:
                                        found = j
                                        break
                                if found:
                                    break
                if found is not None:
                    mapping[dest_idx] = found

            # if mapping exists, transform rows to template order
            if mapping:
                transformed = []
                # add header (template headers)
                transformed.append(template_headers)
                for r in rows[1:]:
                    newrow = [''] * len(template_headers)
                    for dest_idx, src_idx in mapping.items():
                        if src_idx < len(r):
                            newrow[dest_idx] = r[src_idx]
                    transformed.append(newrow)
                out_buf = io.StringIO()
                writer = csv.writer(out_buf)
                for r in transformed:
                    writer.writerow([str(c) for c in r])
                return out_buf.getvalue()

            # fallback: return as-is
            out_buf = io.StringIO()
            writer = csv.writer(out_buf)
            for r in rows:
                writer.writerow([str(c) for c in r])
            return out_buf.getvalue()

        normalized = normalize_ai_csv(csv_text, headers)
        # apply deterministic post-processing to guarantee the CSV matches the template headers
        final_csv = _hardcode_csv_from_ai(normalized, headers, patient_info, doctor_info, synonym_map)
        return final_csv if final_csv and final_csv.strip() else None
    except Exception:
        return None


def _hardcode_csv_from_ai(csv_text: str, template_headers: list[str], patient_info: dict | None, doctor_info: dict | None, synonym_map: dict) -> str:
    """Deterministic transformation: take AI-produced CSV (any header layout), extract key fields using regex/heuristics,
    populate patient/doctor info, normalize booleans/counts, and return a CSV string with EXACT template_headers as the header row."""
    if not csv_text or not csv_text.strip():
        return ''
    lines = [l for l in csv_text.splitlines() if l.strip()]
    try:
        rows = list(csv.reader(lines))
    except Exception:
        return ''
    ai_headers = [h.strip() for h in rows[0]] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    # reverse map for quick synonym lookup
    rev_map = {}
    for std, syns in synonym_map.items():
        for s in syns:
            rev_map[s.lower()] = std

    final_rows = []
    for r in data_rows:
        rtext = ' '.join([str(x) for x in r if x])
        ai_map = {ai_headers[i].strip().lower(): (r[i] if i < len(r) else '') for i in range(len(ai_headers))}
        newrow = [''] * len(template_headers)
        for idx, header in enumerate(template_headers):
            hl = header.strip().lower()
            val = ''
            # direct ai header match
            if hl in ai_map and ai_map[hl].strip():
                val = ai_map[hl].strip()
            else:
                # try match by synonyms / fuzzy
                for k, v in ai_map.items():
                    if not v:
                        continue
                    if hl == k or hl in k or k in hl:
                        val = v.strip()
                        break
                    if k in rev_map and rev_map[k] == hl:
                        val = v.strip()
                        break
            # regex-based extraction when AI didn't provide value
            if not val:
                # medication name
                if 'medikament' in hl or 'medicament' in hl or 'med' in hl:
                    m = re.search(r"([A-Za-zÄÖÜäöüß0-9\-\s]{3,120})\s+(\d+\.?\d*\s*(mg|g|ml|mcg|µg))", rtext, re.I)
                    if m:
                        val = m.group(1).strip()
                    else:
                        m2 = re.search(r"([A-Za-zÄÖÜäöüß0-9\-]{3,60})(?:,|\s|$)", rtext)
                        if m2:
                            val = m2.group(1).strip()
                elif 'anzahl' in hl or 'qty' in hl or 'quantity' in hl:
                    m = re.search(r"\b(\d{1,4})\b", rtext)
                    if m:
                        val = m.group(1)
                elif 'morg' in hl:
                    # prefer explicit markers like 'morgens' or 'morning'
                    val = '1' if re.search(r"\bmorg(en)?\b|\bmorning\b", rtext, re.I) else ''
                elif 'mitt' in hl:
                    val = '1' if re.search(r"\bmitt(ags)?\b|\bnoon\b|\bmidday\b", rtext, re.I) else ''
                elif 'abend' in hl:
                    val = '1' if re.search(r"\babend(s)?\b|\bevening\b", rtext, re.I) else ''
                elif 'täg' in hl or 'daily' in hl:
                    if re.search(r"\b(ja|yes|daily|täglich|taeglich)\b", rtext, re.I):
                        val = 'Ja'
                    elif re.search(r"\b(nein|no)\b", rtext, re.I):
                        val = 'Nein'
                elif any(x in hl for x in ['name', 'kunde', 'patientenangaben']):
                    if patient_info and patient_info.get('name'):
                        val = patient_info.get('name')
                elif 'code' in hl:
                    if patient_info and patient_info.get('code'):
                        val = str(patient_info.get('code'))
                elif any(x in hl for x in ['alter', 'age']):
                    if patient_info and patient_info.get('age'):
                        val = str(patient_info.get('age'))
                elif any(x in hl for x in ['geb', 'birth', 'dob']):
                    if patient_info and patient_info.get('dob'):
                        val = str(patient_info.get('dob'))
                elif 'adresse' in hl or 'address' in hl:
                    if patient_info and patient_info.get('address'):
                        val = patient_info.get('address')
                elif any(x in hl for x in ['behandelnder', 'arzt']):
                    if doctor_info and doctor_info.get('name'):
                        val = doctor_info.get('name')
                elif 'spez' in hl or 'special' in hl:
                    if doctor_info and doctor_info.get('specialty'):
                        val = doctor_info.get('specialty')
            # normalization
            if val:
                # boolean
                if 'täg' in hl or 'daily' in hl:
                    if re.search(r"\b(ja|j|yes|1|daily)\b", str(val), re.I):
                        val = 'Ja'
                    elif re.search(r"\b(nein|no|0)\b", str(val), re.I):
                        val = 'Nein'
                # counts
                if any(x in hl for x in ['morg', 'mitt', 'abend', 'anzahl', 'qty', 'quantity']):
                    m = re.search(r"(\d+)", str(val))
                    if m:
                        val = m.group(1)
            newrow[idx] = val or ''
        final_rows.append(newrow)

    # If AI provided no data rows, try to extract simple lines from the text as a last resort
    if not final_rows:
        # attempt to split csv_text by lines and parse lines with medication-like patterns
        for l in lines:
            if re.search(r"\b\d{1,3}mg\b|\b\d{1,3}g\b|\bAspirin\b|\bParacetamol\b|\b\d+\s*tablets?\b", l, re.I):
                # create a minimal row
                newrow = [''] * len(template_headers)
                for idx, header in enumerate(template_headers):
                    hl = header.strip().lower()
                    if 'medikament' in hl:
                        m = re.search(r"([A-Za-zÄÖÜäöüß0-9\-\s]{3,120})\s+(\d+\s*(mg|g|ml))", l, re.I)
                        if m:
                            newrow[idx] = m.group(1).strip()
                        else:
                            newrow[idx] = l.strip()
                    if any(x in hl for x in ['name', 'kunde']) and patient_info and patient_info.get('name'):
                        newrow[idx] = patient_info.get('name')
                final_rows.append(newrow)

    # build CSV output with exact header order
    out_buf = io.StringIO()
    writer = csv.writer(out_buf)
    writer.writerow([h for h in template_headers])
    for r in final_rows:
        writer.writerow([str(c) for c in r])
    return out_buf.getvalue()


def _ai_extract_patient_id(text: str) -> int | None:
    """Use the AI model to extract a patient id from a short text. Returns integer id or None."""
    if not text or not text.strip():
        return None
    prompt = (
        "You receive an email subject and body. If a patient numeric ID is present, return ONLY the numeric ID (digits) and nothing else. "
        "If no patient id is present, return an empty string. Examples: 'Patient ID: 123' -> '123'. 'Keine ID' -> ''. "
        "Do not return extra commentary or punctuation."
        "\n\nINPUT:\n" + text[:12000]
    )
    try:
        response = client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[
                {"role": "system", "content": "You are a strict data extractor. Output ONLY the digits of the patient id or an empty string."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.0,
            max_tokens=50,
        )
        out = response.choices[0].message.content.strip()
        # try to find digits in the response
        m = re.search(r"(\d{1,10})", out)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        return None
    except Exception:
        return None


def _ai_is_urgent(text: str) -> bool:
    """Return True if text indicates urgency. Use simple heuristics first, fall back to AI that should return 'true' or 'false'."""
    if not text or not text.strip():
        return False
    txt = text.lower()
    # heuristics
    if re.search(r"\b(dr(i|in)gend|sofort|asap|urgent|eilt|immediately|priority|prio)\b", txt, re.I):
        return True
    # fallback to AI
    prompt = (
        "Decide if this email is urgent and requires immediate action. Return ONLY 'true' or 'false' (lowercase), nothing else. "
        "Treat expressions like 'dringend', 'sofort', 'asap', 'urgent' as urgent. Examples: 'Bitte dringend bis morgen' -> 'true'. 'Nur info' -> 'false'.\n\n"
        "INPUT:\n" + text[:6000]
    )
    try:
        response = client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[
                {"role": "system", "content": "You are a concise classifier. Output only 'true' or 'false'."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.0,
            max_tokens=10,
        )
        out = response.choices[0].message.content.strip().lower()
        if 'true' in out or out.strip() == 'true' or re.search(r"\b(ja|yes)\b", out):
            return True
        return False
    except Exception:
        return False


def _parse_date_string(s: str) -> str | None:
    """Try to parse a date-like string into ISO yyyy-mm-dd. Returns string or None."""
    if not s or not s.strip():
        return None
    s = s.strip()
    # try a set of common formats
    fmts = ["%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"]
    for f in fmts:
        try:
            import datetime
            dt = datetime.datetime.strptime(s, f)
            return dt.date().isoformat()
        except Exception:
            continue
    # try to extract dd.mm.yyyy from inside text
    m = re.search(r"(\d{1,2}[\.\-/]\d{1,2}[\.\-/]\d{2,4})", s)
    if m:
        for f in fmts:
            try:
                import datetime
                dt = datetime.datetime.strptime(m.group(1), f)
                return dt.date().isoformat()
            except Exception:
                continue
    return None


def _ai_extract_done_by(text: str) -> str | None:
    """Try to find a 'due date' in text. Use regex heuristics first, fallback to AI which should return ISO date or empty string."""
    if not text or not text.strip():
        return None
    combined = text
    # heuristics: look for 'bis <date>' or 'by <date>' or plain dates
    m = re.search(r"\b(?:bis|by)[:\s]*([\d\.\-/ ]{6,20})", combined, re.I)
    if m:
        parsed = _parse_date_string(m.group(1))
        if parsed:
            return parsed
    # plain date patterns
    m2 = re.search(r"(\d{1,2}[\.\-/]\d{1,2}[\.\-/]\d{2,4})", combined)
    if m2:
        parsed = _parse_date_string(m2.group(1))
        if parsed:
            return parsed
    # fallback to AI
    prompt = (
        "Find the deadline or due date in this email. If found, return ONLY the date in ISO format YYYY-MM-DD. "
        "If no date or deadline is present, return an empty string. Examples: 'bis 12.01.2026' -> '2026-01-12'.\n\n"
        "INPUT:\n" + text[:8000]
    )
    try:
        response = client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[
                {"role": "system", "content": "You are a precise date extractor. Output the date in YYYY-MM-DD or empty string."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.0,
            max_tokens=50,
        )
        out = response.choices[0].message.content.strip()
        # try to find ISO date
        m_iso = re.search(r"(20\d{2}-\d{2}-\d{2})", out)
        if m_iso:
            return m_iso.group(1)
        # try to parse returned text
        parsed = _parse_date_string(out)
        if parsed:
            return parsed
        return None
    except Exception:
        return None


@app.post("/emails/analyze_eml")
async def analyze_eml(file: UploadFile = File(...)):
    """Analyze an uploaded .eml and return a breakdown: { 'from', 'subject', 'text', 'patient_id' }.
    The route first uses deterministic heuristics to find patient id (numeric id or 5-digit code), then falls back to AI extraction.
    """
    if not file:
        raise HTTPException(status_code=400, detail="Provide an .eml file")
    fname = (file.filename or "uploaded").lower()
    if not (fname.endswith('.eml') or file.content_type == 'message/rfc822'):
        raise HTTPException(status_code=400, detail="File must be a .eml file")
    data = await file.read()
    from email.parser import BytesParser
    parser = BytesParser(policy=policy.default)
    try:
        msg = parser.parsebytes(data)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid .eml file")

    subject = msg.get('Subject') or ''
    from_addr = msg.get('From') or ''
    text = _get_plain_body_from_message(msg)

    # heuristics: numeric patient id in subject/body
    pid = None
    id_m = re.search(r"\b(?:patient|id)\s*[:#-]?\s*(\d{1,10})\b", (subject + "\n" + text), re.I)
    if id_m:
        try:
            pid = int(id_m.group(1))
        except Exception:
            pid = None
    # try 5-digit patient code and resolve to id
    if pid is None:
        code_m = re.search(r"\b(\d{5})\b", (subject + "\n" + text))
        if code_m:
            code = code_m.group(1)
            # resolve via DB
            try:
                conn = sqlite3.connect(str(DB_PATH))
                cur = conn.cursor()
                cur.execute("SELECT id FROM patients WHERE code = ?", (code,))
                r = cur.fetchone()
                if r:
                    pid = int(r[0])
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

    # try name match
    if pid is None:
        try:
            conn = sqlite3.connect(str(DB_PATH))
            cur = conn.cursor()
            cur.execute("SELECT id, name FROM patients")
            for r in cur.fetchall():
                if r[1] and r[1].lower() in (subject + " " + text).lower():
                    pid = int(r[0])
                    break
        finally:
            try:
                conn.close()
            except Exception:
                pass

    # fallback to AI extraction
    if pid is None:
        ai_pid = _ai_extract_patient_id(f"Subject:\n{subject}\n\nBody:\n{text}")
        if ai_pid:
            # if ai returned a 5-digit code, resolve to id if possible
            if len(str(ai_pid)) == 5:
                try:
                    conn = sqlite3.connect(str(DB_PATH))
                    cur = conn.cursor()
                    cur.execute("SELECT id FROM patients WHERE code = ?", (str(ai_pid),))
                    r = cur.fetchone()
                    if r:
                        pid = int(r[0])
                    else:
                        # treat ai_pid as direct id
                        pid = int(ai_pid)
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            else:
                pid = ai_pid

    # detect urgency and due date (heuristics first, then AI)
    urgent = _ai_is_urgent(subject + "\n" + text)
    done_by = _ai_extract_done_by(subject + "\n" + text)
    # convert ISO string to a date object for JSON date output
    done_by_obj = None
    if done_by:
        try:
            done_by_obj = datetime.date.fromisoformat(done_by)
        except Exception:
            # if parsing fails, return original string
            done_by_obj = done_by

    return {"from": from_addr, "subject": subject, "text": text, "patient_id": pid, "urgent": urgent, "done_by": done_by_obj}

#create/init the database
@app.post("/init_db")
async def init_db():
    """Create sample `doktores` table and populate five entries."""
    print("Starting init_db...")
    import time
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Add timeout to handle 'database is locked' errors
            conn = sqlite3.connect(str(DB_PATH), timeout=30)
            cur = conn.cursor()
            
            print("Initializing doktores table...")
            cur.execute("CREATE TABLE IF NOT EXISTS doktores (id INTEGER PRIMARY KEY, name TEXT NOT NULL, specialty TEXT NOT NULL, email TEXT)")
            cur.execute("DELETE FROM doktores")
            sample_doctors = [
                (1, "Dr. Anna Müller", "Cardiology", "anna.mueller@clinic.ch"),
                (2, "Dr. Lukas Meier", "Dermatology", "lukas.meier@clinic.ch"),
                (3, "Dr. Martina Schmid", "Pediatrics", "martina.schmid@clinic.ch"),
                (4, "Dr. Stefan Keller", "Neurology", "stefan.keller@clinic.ch"),
                (5, "Dr. Sophie Weber", "Orthopedics", "sophie.weber@clinic.ch"),
            ]
            cur.executemany("INSERT INTO doktores (id, name, specialty, email) VALUES (?, ?, ?, ?)", sample_doctors)

            print("Initializing patients table...")
            cur.execute("CREATE TABLE IF NOT EXISTS patients (id INTEGER PRIMARY KEY, name TEXT NOT NULL, address TEXT, age INTEGER, conditions TEXT)")
            cur.execute("DELETE FROM patients")
            sample_patients = [
                (1, "Hans Müller", "Bahnhofstrasse 10, Zürich", 54, "hypertension, asthma"),
                (2, "Monika Steiner", "Gartenweg 4, Bern", 68, "allergies"),
                (3, "Marco Rossi", "Via Nassa 12, Lugano", 52, "diabetes"),
                (4, "Aisha Huber", "Wiesenstrasse 88, Lausanne", 89, "heart disease"),
                (5, "Pierre Dubois", "Rue de la Paix 12, Genève", 65, "arthritis, hypertension"),
                (6, "Olivia Meier", "Birkenweg 210, Luzern", 71, "migraine"),
                (7, "Noah Schmid", "Lindenweg 3, St. Gallen", 76, "asthma, high cholesterol"),
                (8, "Emma Weber", "Hafengasse 56, Basel", 93, "diabetes, arthritis"),
                (9, "Marta Novak", "Kirchgasse 9, Winterthur", 58, "cardiac arrhythmia"),
                (10, "Samuel Frei", "Tuchlaedeli 77, Thun", 76, "high cholesterol"),
            ]
            cur.executemany("INSERT INTO patients (id, name, address, age, conditions) VALUES (?, ?, ?, ?, ?)", sample_patients)

            print("Initializing users table...")
            cur.execute("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, role TEXT NOT NULL, password TEXT NOT NULL)")
            cur.execute("DELETE FROM users")
            sample_users = [
                (1, "doctor", "doc1pass"), (2, "doctor", "doc2pass"), (3, "doctor", "doc3pass"), (4, "doctor", "doc4pass"), (5, "doctor", "doc5pass"),
                (1001, "nurse", "nurse1pass"), (1002, "nurse", "nurse2pass"), (1003, "nurse", "nurse3pass"), (1004, "nurse", "nurse4pass"), (1005, "nurse", "nurse5pass"),
                (2001, "admin", "admin1pass"), (2002, "admin", "admin2pass"),
            ]
            cur.executemany("INSERT INTO users (id, role, password) VALUES (?, ?, ?)", sample_users)

            print("Initializing documents and emails tables...")
            cur.execute("CREATE TABLE IF NOT EXISTS documents (id INTEGER PRIMARY KEY, patient_id INTEGER NOT NULL, filename TEXT NOT NULL, content BLOB, size INTEGER NOT NULL, attachment_path TEXT, uploaded_at TEXT DEFAULT (datetime('now')), FOREIGN KEY(patient_id) REFERENCES patients(id))")
            cur.execute("DELETE FROM documents")
            cur.execute("CREATE TABLE IF NOT EXISTS emails (id INTEGER PRIMARY KEY, patient_id INTEGER, subject TEXT NOT NULL, body TEXT, attachment_filename TEXT, attachment_content BLOB, attachment_path TEXT, attachment_size INTEGER, created_at TEXT DEFAULT (datetime('now')), FOREIGN KEY(patient_id) REFERENCES patients(id))")
            cur.execute("DELETE FROM emails")

            print("Initializing chat_messages table...")
            cur.execute("DROP TABLE IF EXISTS chat_messages")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS chat_messages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    client_id TEXT NOT NULL,
                    sender_id TEXT NOT NULL,
                    recipient_id TEXT NOT NULL,
                    message_type TEXT NOT NULL DEFAULT 'message',
                    content TEXT,
                    is_important BOOLEAN DEFAULT 0,
                    due_date TEXT,
                    status TEXT DEFAULT 'pending',
                    file_path TEXT,
                    file_name TEXT,
                    created_at TEXT DEFAULT (datetime('now'))
                )
            """)
            cur.execute("DELETE FROM chat_messages")

            print("Setting up storage directories...")
            STORAGE_DIR.mkdir(parents=True, exist_ok=True)
            STORAGE_DOCS_DIR.mkdir(parents=True, exist_ok=True)
            STORAGE_EMAILS_DIR.mkdir(parents=True, exist_ok=True)

            print("Ensuring patient codes...")
            cur.execute("PRAGMA table_info(patients)")
            if 'code' not in [r[1] for r in cur.fetchall()]:
                cur.execute("ALTER TABLE patients ADD COLUMN code TEXT")
                cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_patients_code ON patients(code)")

            cur.execute("SELECT id FROM patients")
            pids = [r[0] for r in cur.fetchall()]
            existing_codes = set()
            for pid in pids:
                new_code = _generate_code(5)
                while new_code in existing_codes:
                    new_code = _generate_code(5)
                existing_codes.add(new_code)
                cur.execute("UPDATE patients SET code = ? WHERE id = ?", (new_code, pid))

            print("Initializing patient_doctors table...")
            cur.execute("CREATE TABLE IF NOT EXISTS patient_doctors (patient_id INTEGER NOT NULL, doctor_id INTEGER NOT NULL, PRIMARY KEY (patient_id, doctor_id), FOREIGN KEY(patient_id) REFERENCES patients(id), FOREIGN KEY(doctor_id) REFERENCES doktores(id))")
            cur.execute("DELETE FROM patient_doctors")

            print("Populating patient_doctors assignments...")
            cur.execute("SELECT id FROM doktores")
            dr_ids = [r[0] for r in cur.fetchall()]
            if dr_ids:
                assignments = []
                for pid in pids:
                    num = random.randint(2, min(3, len(dr_ids)))
                    for d_id in random.sample(dr_ids, num):
                        assignments.append((pid, d_id))
                cur.executemany("INSERT INTO patient_doctors (patient_id, doctor_id) VALUES (?, ?)", assignments)

            print("Initializing nurses table...")
            cur.execute("CREATE TABLE IF NOT EXISTS nurses (id INTEGER PRIMARY KEY, name TEXT NOT NULL)")
            cur.execute("DELETE FROM nurses")
            sample_nurses = [
                (1001, "Elena Fischer"),
                (1002, "Marcus Weber"),
                (1003, "Julia Baumgartner"),
                (1004, "Thomas Herzog"),
                (1005, "Sarah Schneider"),
            ]
            cur.executemany("INSERT INTO nurses (id, name) VALUES (?, ?)", sample_nurses)

            conn.commit()
            conn.close()
            print("init_db success.")
            return {"ok": True, "db_path": str(DB_PATH)}
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e) and attempt < max_retries - 1:
                print(f"Database locked, retrying (attempt {attempt + 1})...")
                time.sleep(1)
                continue
            print(f"init_db OperationalError: {e}")
            raise HTTPException(status_code=500, detail=f"Database error: {e}")
        except Exception as e:
            print(f"init_db ERROR: {e}")
            raise HTTPException(status_code=500, detail=str(e))

#-- Fetch data endpoints ---
@app.get("/doktores_db")
async def doktores_db():
    """Fetch doktores from the SQLite database file `spitex.db`."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, name, specialty, email FROM doktores ORDER BY id")
        rows = cur.fetchall()
        result = [dict(r) for r in rows]
        return {"doktores": result}
    finally:
        conn.close()

#-- Fetch data endpoints ---
@app.get("/patients_db")
async def patients_db():
    """Fetch patients from the DB, including their unique code and assigned doctors."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, code, name, address, age, conditions FROM patients ORDER BY id")
        rows = cur.fetchall()
        result = [dict(r) for r in rows]
        
        # for each patient, fetch assigned doctors
        for p in result:
            cur.execute("""
                SELECT d.id, d.name, d.specialty, d.email 
                FROM doktores d
                JOIN patient_doctors pd ON d.id = pd.doctor_id
                WHERE pd.patient_id = ?
            """, (p['id'],))
            p['doctors'] = [dict(r) for r in cur.fetchall()]
            
        return {"patients": result}
    finally:
        conn.close()

@app.get("/patient_doctors_db")
async def patient_doctors_db():
    """Fetch all patient-doctor assignments."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT pd.patient_id, p.name as patient_name, pd.doctor_id, d.name as doctor_name
            FROM patient_doctors pd
            JOIN patients p ON pd.patient_id = p.id
            JOIN doktores d ON pd.doctor_id = d.id
        """)
        rows = cur.fetchall()
        return {"assignments": [dict(r) for r in rows]}
    finally:
        conn.close()

@app.get("/nurses_db")
async def nurses_db():
    """Fetch nurses from the DB."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, name FROM nurses ORDER BY id")
        rows = cur.fetchall()
        result = [dict(r) for r in rows]
        return {"nurses": result}
    finally:
        conn.close()

#-- Fetch users/login entries endpoint ---
@app.get("/users_db")
async def users_db():
    """Fetch users/login entries from the DB."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, role, password FROM users ORDER BY id")
        rows = cur.fetchall()
        result = [dict(r) for r in rows]
        return {"users": result}
    finally:
        conn.close()

@app.get("/messages_db")
async def messages_db(client_id: str | None = None):
    """Fetch chat messages from SQLite, optionally filtered by client_id."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        if client_id:
            cur.execute("SELECT * FROM chat_messages WHERE client_id = ? ORDER BY created_at ASC", (client_id,))
        else:
            cur.execute("SELECT * FROM chat_messages ORDER BY created_at ASC")
        rows = cur.fetchall()
        return {"messages": [dict(r) for r in rows]}
    finally:
        conn.close()

@app.post("/messages_db")
async def save_message_db(msg: MessageSchema):
    """Save a new chat message to SQLite."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO chat_messages 
            (client_id, sender_id, recipient_id, content, message_type, is_important, due_date, status, file_path, file_name)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (msg.client_id, msg.sender_id, msg.recipient_id, msg.content, msg.message_type, 1 if msg.is_important else 0, msg.due_date, msg.status, msg.file_path, msg.file_name))
        new_id = cur.lastrowid
        conn.commit()
        
        # Fetch the newly created message to return it
        cur.row_factory = sqlite3.Row
        cur.execute("SELECT * FROM chat_messages WHERE id = ?", (new_id,))
        row = cur.fetchone()
        return dict(row)
    finally:
        conn.close()

#-- Example PDFs generation endpoint ---
@app.post("/examples/init")
async def examples_init():
    """Create the `examples/pdfs` directory and generate sample PDFs (requires reportlab)."""
    EXAMPLES_DIR.mkdir(parents=True, exist_ok=True)
    if canvas is None:
        raise HTTPException(status_code=500, detail="PDF generation requires 'reportlab'. Install with: pip install reportlab")

    # generate 3 simple PDFs
    for i in range(1, 4):
        path = EXAMPLES_DIR / f"example_{i}.pdf"
        buf_path = str(path)
        c = canvas.Canvas(buf_path)
        c.setTitle(f"Beispiel {i}")
        c.setFont("Helvetica-Bold", 20)
        c.drawString(72, 750, f"Beispiel PDF {i}")
        c.setFont("Helvetica", 12)
        c.drawString(72, 720, "Dies ist ein Beispiel-PDF, das für Demonstrationszwecke erstellt wurde.")
        c.save()

    return {"ok": True, "examples_dir": str(EXAMPLES_DIR)}

#-- Document management endpoints ---
@app.get("/patients/{patient_identifier}/documents")
async def list_patient_documents(patient_identifier: str):
    """List documents (PDFs) for a given patient identifier (id or code)."""
    pid, pcode, pname = _resolve_patient(patient_identifier)
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT filename, size, attachment_path, uploaded_at FROM documents WHERE patient_id = ? ORDER BY uploaded_at DESC", (pid,))
        rows = cur.fetchall()
        result = [dict(r) for r in rows]
        # prefer to show patient code as folder name
        for r in result:
            if r.get('attachment_path'):
                # normalize to show storage path with code if possible
                r['attachment_path'] = r['attachment_path']
        return {"patient": {"id": pid, "code": pcode, "name": pname}, "documents": result}
    finally:
        conn.close()

#-- Upload document endpoints ---
@app.post("/patients/{patient_identifier}/documents/upload")
async def upload_patient_document(patient_identifier: str, file: UploadFile = File(...)):
    """Upload a PDF and attach it to a patient (stored as file on disk and BLOB in DB for backward compatibility)."""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only .pdf files are accepted")
    content = await file.read()
    size = len(content)

    pid, pcode, pname = _resolve_patient(patient_identifier)

    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        safe_filename = os.path.basename(file.filename)
        # ensure storage directory exists (use patient code for foldering)
        dest_dir = STORAGE_DOCS_DIR / (pcode or str(pid))
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_path = dest_dir / safe_filename
        # avoid overwriting: add numeric suffix if exists
        counter = 1
        base, ext = os.path.splitext(safe_filename)
        while dest_path.exists():
            safe_filename = f"{base}_{counter}{ext}"
            dest_path = dest_dir / safe_filename
            counter += 1
        with dest_path.open('wb') as out_f:
            out_f.write(content)
        # store both BLOB (for backward compatibility) and path
        cur.execute("INSERT INTO documents (patient_id, filename, content, size, attachment_path) VALUES (?, ?, ?, ?, ?)",
                    (pid, safe_filename, sqlite3.Binary(content), size, str(dest_path)))
        conn.commit()
        return {"ok": True, "filename": safe_filename, "size": size, "path": str(dest_path)}
    finally:
        conn.close()

#-- Download document endpoint ---
@app.get("/patients/{patient_identifier}/documents/{filename}")
async def download_patient_document(patient_identifier: str, filename: str):
    """Download a named PDF attached to a patient (identifier can be numeric id or code)."""
    pid, pcode, pname = _resolve_patient(patient_identifier)
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        cur.execute("SELECT content, filename, attachment_path FROM documents WHERE patient_id = ? AND filename = ? LIMIT 1", (pid, filename))
        row = cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Document not found")
        # prefer file on disk if attachment_path exists (path may include code in its path)
        att_path = row["attachment_path"] if "attachment_path" in row.keys() else None
        if att_path:
            path = Path(att_path)
            if path.exists():
                return FileResponse(path, media_type="application/pdf", filename=row["filename"])
        # fallback to BLOB
        content = row["content"]
        if content is None:
            raise HTTPException(status_code=404, detail="Document content not available")
        disp = f'attachment; filename="{row["filename"]}"'
        return StreamingResponse(io.BytesIO(content), media_type="application/pdf", headers={"Content-Disposition": disp})
    finally:
        conn.close()

#-- Example PDFs endpoints ---
@app.post("/examples/populate_documents")
async def populate_documents(min_per_patient: int = 2, max_per_patient: int = 3):
    """Populate the `documents` table: attach 2..3 example PDFs to each patient.

    - If there are fewer than `min_per_patient` documents for a patient, this endpoint
      uploads additional example PDFs (from `examples/pdfs`) until the patient has
      at least `min_per_patient` docs, up to `max_per_patient`.
    - Returns summary of insertions.
    """
    # ensure example files exist (uses the same generator as /examples/init)
    await examples_init()
    files = [f for f in sorted(EXAMPLES_DIR.iterdir()) if f.is_file()]
    if not files:
        raise HTTPException(status_code=500, detail="No example PDFs available to populate documents")

    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, code FROM patients")
        patients = [r for r in cur.fetchall()]
        inserted = 0
        patients_updated = 0
        for pid, pcode in patients:
            cur.execute("SELECT COUNT(*) FROM documents WHERE patient_id = ?", (pid,))
            (count,) = cur.fetchone()
            if count >= min_per_patient:
                continue
            target = random.randint(min_per_patient, max_per_patient)
            to_add = max(0, target - count)
            added_for_patient = 0
            # pick files (allow duplicates across patients)
            for i in range(to_add):
                src = files[i % len(files)]
                with src.open('rb') as f:
                    content = f.read()
                # write to storage on disk (use code folder when available)
                folder = pcode or str(pid)
                dest_dir = STORAGE_DOCS_DIR / folder
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest_path = dest_dir / src.name
                counter = 1
                base, ext = os.path.splitext(src.name)
                while dest_path.exists():
                    dest_path = dest_dir / f"{base}_{counter}{ext}"
                    counter += 1
                with dest_path.open('wb') as out_f:
                    out_f.write(content)
                cur.execute(
                    "INSERT INTO documents (patient_id, filename, content, size, attachment_path) VALUES (?, ?, ?, ?, ?)",
                    (pid, dest_path.name, sqlite3.Binary(content), len(content), str(dest_path)))
                inserted += 1
                added_for_patient += 1
            if added_for_patient:
                patients_updated += 1
        conn.commit()
        return {"ok": True, "patients_updated": patients_updated, "inserted": inserted}
    finally:
        conn.close()

#-- Email example creation ---
@app.post("/examples/populate_emails")
async def populate_emails(count: int = 5):
    """Create `count` example emails (default 5), each with a PDF attachment.

    Attachments are taken from `examples/pdfs` and emails are assigned to random patients.

    Important: emails are stored as .eml files on disk only and NOT inserted into the database.
    """
    # ensure example files exist
    await examples_init()
    files = [f for f in sorted(EXAMPLES_DIR.iterdir()) if f.is_file()]
    if not files:
        raise HTTPException(status_code=500, detail="No example PDFs available to attach")

    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, code FROM patients")
        patients = [r for r in cur.fetchall()]
        if not patients:
            raise HTTPException(status_code=500, detail="No patients in database to assign emails to")
        created = []
        # always generate exactly 5 example emails as requested
        count = 5
        topics = [
            "Medikamentenumstellung",
            "Rezeptnachbestellung",
            "Überprüfung Laborergebnisse",
            "Frage zur Medikation",
            "Follow-up: Medikation",
            "Therapieanpassung",
            "Anfrage: Klärung Dosierung",
        ]
        for i in range(count):
            pid, pcode = random.choice(patients)
            src = files[i % len(files)]
            with src.open('rb') as f:
                content = f.read()

            # write attachment (pdf) to storage on disk for that patient (use code folder)
            folder = pcode or str(pid)
            dest_dir = STORAGE_EMAILS_DIR / folder
            dest_dir.mkdir(parents=True, exist_ok=True)
            base, ext = os.path.splitext(src.name)
            pdf_dest = dest_dir / src.name
            counter = 1
            while pdf_dest.exists():
                pdf_dest = dest_dir / f"{base}_{counter}{ext}"
                counter += 1
            with pdf_dest.open('wb') as out_f:
                out_f.write(content)

            # pick a doctor to be the sender (must be from doktores table)
            cur.execute("SELECT id, name, email FROM doktores")
            doktores = cur.fetchall()
            if doktores:
                dr = random.choice(doktores)
                dr_id, dr_name, dr_email = dr[0], dr[1], dr[2] if len(dr) > 2 else None
                if not dr_email:
                    # craft a plausible email if missing
                    dr_email = f"{dr_name.split()[-1].lower()}@example.local"
            else:
                dr_id, dr_name, dr_email = None, 'Dr. Example', 'doctor@example.local'

            # get patient name and code
            cur.execute("SELECT name, code FROM patients WHERE id = ?", (pid,))
            p_row = cur.fetchone()
            patient_name = p_row[0] if p_row else f"patient-{pid}"
            patient_code = p_row[1] if p_row and p_row[1] else None

            # ensure we have a 5-digit patient code; if not, generate one and persist it
            if not (patient_code and re.fullmatch(r"\d{5}", str(patient_code))):
                # fetch existing codes to avoid collision
                cur.execute("SELECT code FROM patients WHERE code IS NOT NULL")
                existing = {r[0] for r in cur.fetchall() if r[0]}
                new_code = None
                for _ in range(200):
                    cand = _generate_code(5)
                    if cand not in existing:
                        new_code = cand
                        break
                if new_code is None:
                    # as a last resort, zero-pad the numeric pid
                    new_code = str(pid).zfill(5)[-5:]
                try:
                    cur.execute("UPDATE patients SET code = ? WHERE id = ?", (new_code, pid))
                    conn.commit()
                    patient_code = new_code
                except Exception:
                    # if update fails, still use the candidate
                    patient_code = new_code

            # sent date (recent past)
            sent_dt = datetime.datetime.now() - datetime.timedelta(days=random.randint(0, 7))

            # always include a requested-by date (random 1-14 days ahead)
            due_dt = (datetime.date.today() + datetime.timedelta(days=random.randint(1, 14)))
            deadline_str = due_dt.strftime('%d.%m.%Y')

            # 50% chance to mark as urgent
            is_urgent = random.random() < 0.5

            # subject: medical topic + patient name or code
            topic = random.choice(topics)
            subject_patient = patient_code if random.random() < 0.5 and patient_code else patient_name
            subj_prefix = "DRINGEND: " if is_urgent else ""
            subject = f"{subj_prefix}{topic} - {subject_patient}"

            # filler text
            fillers = [
                "Bitte beachten Sie die Hinweise in der Anlage.",
                "Nur zur Info, keine Aktion erforderlich.",
                "Falls Rückfragen, melden Sie sich bitte.",
                "Dies ist ein Platzhaltertext für medizinische Dokumentation.",
                "Zusätzliche Anmerkungen: Dies ist ein Beispiel.",
                "Referenzcode: " + uuid.uuid4().hex[:6]
            ]
            extra = "\n".join(random.sample(fillers, k=random.randint(1, 2)))

            body_lines = [
                f"Arzt: {dr_name} <{dr_email}>",
                f"Patientenname: {patient_name}",
                f"Patienten-ID: {patient_code or pid}",
            ]
            body_lines.append("")
            body_lines.append(f"{topic} bezüglich des oben genannten Patienten.")
            body_lines.append(f"Angefordert bis: {deadline_str}")
            if is_urgent:
                body_lines.append("Bitte dringend bearbeiten.")
            body_lines.append("")
            body_lines.append(extra)
            body = "\n".join(body_lines)

            msg = EmailMessage()
            msg["From"] = dr_email
            msg["To"] = INBOX_ADDRESS
            msg["Subject"] = subject
            msg["Date"] = format_datetime(sent_dt)
            msg.set_content(body)
            msg.add_attachment(content, maintype="application", subtype="pdf", filename=pdf_dest.name)
            eml_bytes = msg.as_bytes(policy=policy.default)

            # write .eml to disk (unique name)
            eml_dest = dest_dir / f"email_{uuid.uuid4().hex}.eml"
            with eml_dest.open('wb') as ef:
                ef.write(eml_bytes)

            created.append({
                # `patient_id` must be the 5-digit patient code (string) only
                "patient_id": str(patient_code),
                "patient_code": pcode,
                "patient_identifier": (patient_code or pcode or pid),
                "eml": str(eml_dest),
                "attachment": str(pdf_dest),
                "subject": subject,
                "date": sent_dt.isoformat(),
                "urgent": bool(is_urgent),
                "requested_by": deadline_str
            })
        return {"ok": True, "created": created}
    finally:
        conn.close()

@app.post("/emails/upload")
async def upload_eml(file: UploadFile = File(...)):
    """Upload an .eml file. Attempt to parse patient id (from Subject or body) and
    store the .eml in storage/emails/{patient_id}/ or storage/emails/unsorted when unknown.
    Returns metadata about the stored file and any parsed fields.
    """
    filename = file.filename or "uploaded.eml"
    if not filename.lower().endswith('.eml') and file.content_type not in ("message/rfc822", "application/octet-stream"):
        raise HTTPException(status_code=400, detail="Only .eml files are accepted")
    content = await file.read()

    from email.parser import BytesParser
    parser = BytesParser(policy=policy.default)
    patient_id = None
    subject = None
    from_addr = None
    to_addr = None
    date = None
    try:
        msg = parser.parsebytes(content)
        subject = msg.get('Subject')
        from_addr = msg.get('From')
        to_addr = msg.get('To')
        date = msg.get('Date')
        # try to find patient id in subject first
        m = re.search(r"Patient\s*[:]?\s*(\d+)", subject or "")
        if m:
            patient_id = int(m.group(1))
        else:
            # search in plain text body
            body = None
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type().startswith('text/'):
                        try:
                            body = part.get_content()
                            break
                        except Exception:
                            continue
            else:
                try:
                    body = msg.get_content()
                except Exception:
                    body = None
            if body:
                m2 = re.search(r"id\s*[:]?\s*(\d+)", body)
                if m2:
                    patient_id = int(m2.group(1))

        # if no numeric id found, try to find a 5-digit patient code in subject or body
        if patient_id is None:
            code_m = re.search(r"\b(\d{5})\b", subject or "") or (body and re.search(r"\b(\d{5})\b", body))
            if code_m:
                code = code_m.group(1)
                try:
                    conn2 = sqlite3.connect(str(DB_PATH))
                    cur2 = conn2.cursor()
                    cur2.execute("SELECT id FROM patients WHERE code = ?", (code,))
                    r = cur2.fetchone()
                    if r:
                        patient_id = r[0]
                except Exception:
                    pass
                finally:
                    try:
                        conn2.close()
                    except Exception:
                        pass
    except Exception:
        msg = None

    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        if patient_id is not None:
            # try to use patient code if available
            cur.execute("SELECT code FROM patients WHERE id = ?", (patient_id,))
            row = cur.fetchone()
            folder = row[0] if row and row[0] else str(patient_id)
            dest_dir = STORAGE_EMAILS_DIR / folder
        else:
            dest_dir = STORAGE_EMAILS_DIR / "unsorted"
        dest_dir.mkdir(parents=True, exist_ok=True)
    finally:
        conn.close()

    base = Path(filename).stem
    ext = Path(filename).suffix or '.eml'
    dest_path = dest_dir / filename
    counter = 1
    while dest_path.exists():
        dest_path = dest_dir / f"{base}_{counter}{ext}"
        counter += 1
    with dest_path.open('wb') as out_f:
        out_f.write(content)

    # attempt to also return the patient code if we stored under a code folder
    parent = dest_path.parent
    patient_code = parent.name if parent.name != 'unsorted' else None
    return {
        "ok": True,
        "filename": dest_path.name,
        "path": str(dest_path),
        "patient_id": patient_id,
        "patient_code": patient_code,
        "subject": subject,
        "from": from_addr,
        "to": to_addr,
        "date": date,
        "size": dest_path.stat().st_size
    }


@app.post("/emails/extract_and_fill_template")
async def extract_and_fill_template(file: UploadFile | None = File(None), text: str | None = None, as_zip: bool = True, template: str | None = None):
    """Accept an uploaded .eml, CSV, PDF, or plain text and try to extract any medication prescriptions.
    If found, fill the Excel template `Medikamentenliste_Arzt.xlsx` and return the filled workbook.
    - file: optional uploaded file (.eml, .csv, .pdf, .txt)
    - text: optional raw text input
    - as_zip: if true, return a zip containing the filled xlsx and original CSV (if any)
    - template: optional path to the workbook template; defaults to examples/Medikamentenliste_Arzt.xlsx
    """
    if not file and not text:
        raise HTTPException(status_code=400, detail="Provide a file or text input to scan")

    agg_text_parts = []
    original_csv_bytes = None
    original_csv_name = None

    if file:
        fname = (file.filename or "uploaded").lower()
        data = await file.read()
        if fname.endswith('.eml') or file.content_type == 'message/rfc822':
            from email.parser import BytesParser
            parser = BytesParser(policy=policy.default)
            try:
                msg = parser.parsebytes(data)
            except Exception:
                raise HTTPException(status_code=400, detail='Invalid .eml file')
            # body
            agg_text_parts.append(_get_plain_body_from_message(msg))
            # attachments
            for part in msg.iter_attachments():
                pname = part.get_filename() or 'attachment'
                try:
                    payload = part.get_content()
                except Exception:
                    try:
                        payload = part.get_payload(decode=True)
                    except Exception:
                        payload = None
                if isinstance(payload, str):
                    agg_text_parts.append(payload)
                    if pname.lower().endswith('.csv'):
                        original_csv_bytes = payload.encode('utf-8')
                        original_csv_name = pname
                elif isinstance(payload, (bytes, bytearray)):
                    if pname.lower().endswith('.csv') or part.get_content_type() == 'text/csv':
                        try:
                            txt = payload.decode('utf-8', errors='replace')
                            agg_text_parts.append(txt)
                            original_csv_bytes = txt.encode('utf-8')
                            original_csv_name = pname
                        except Exception:
                            pass
                    elif pname.lower().endswith('.pdf') or part.get_content_type() == 'application/pdf':
                        pdftext = _extract_text_from_pdf_bytes(payload)
                        if pdftext:
                            agg_text_parts.append(pdftext)
        elif fname.endswith('.csv') or file.content_type == 'text/csv':
            # treat directly as CSV attachment
            original_csv_bytes = data
            original_csv_name = Path(file.filename).name
            try:
                s = data.decode('utf-8', errors='replace')
            except Exception:
                s = ''
            agg_text_parts.append(s)
        elif fname.endswith('.pdf') or file.content_type == 'application/pdf':
            pdftext = _extract_text_from_pdf_bytes(data)
            if pdftext:
                agg_text_parts.append(pdftext)
        else:
            try:
                s = data.decode('utf-8', errors='replace')
                agg_text_parts.append(s)
            except Exception:
                pass

    if text:
        agg_text_parts.append(text)

    agg_text = "\n\n".join([p for p in agg_text_parts if p])

    # determine template path
    tpl_path = Path(template) if template else Path(__file__).parent / 'examples' / 'Medikamentenliste_Arzt.xlsx'
    if not tpl_path.exists():
        raise HTTPException(status_code=400, detail=f"Template not found: {tpl_path}")
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is required to fill Excel templates. Install with: pip install openpyxl")

    # read headers from template (first non-empty row)
    wb = openpyxl.load_workbook(str(tpl_path))
    ws = wb.active
    header_row = None
    for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if any([c is not None and str(c).strip() != '' for c in r]):
            header_row = [str(c) if c is not None else '' for c in r]
            break
    if not header_row:
        raise HTTPException(status_code=500, detail="Could not determine header row in template")

    # attempt to detect patient in the text: id, code, name, address, age, dob
    patient_info = {"id": None, "code": None, "name": None, "address": None, "age": None, "dob": None}
    id_m = re.search(r"\b(?:patient|id)\s*[:#-]?\s*(\d+)\b", agg_text, re.I)
    if id_m:
        patient_info['id'] = int(id_m.group(1))
    else:
        code_m = re.search(r"\b(\d{5})\b", agg_text)
        if code_m:
            patient_info['code'] = code_m.group(1)
    # try to detect simple address/age/dob patterns
    addr_m = re.search(r"(Address|Adresse)[:\s]+([^\n\r]+)", agg_text, re.I)
    if addr_m:
        patient_info['address'] = addr_m.group(2).strip()
    age_m = re.search(r"(age|alter)[:\s]+(\d{1,3})", agg_text, re.I)
    if age_m:
        patient_info['age'] = age_m.group(2)
    dob_m = re.search(r"(geburtstag|geburtsdatum|dob)[:\s]+([0-3]?\d[\.\-/][01]?\d[\.\-/][12][0-9]{3})", agg_text, re.I)
    if dob_m:
        patient_info['dob'] = dob_m.group(2)

    # try name matching against patients table when no id/code found
    if not patient_info['id'] and not patient_info['code']:
        try:
            conn = sqlite3.connect(str(DB_PATH))
            cur = conn.cursor()
            cur.execute("SELECT id, name, code, address, age FROM patients")
            for pid, name, code, addr, age in cur.fetchall():
                if name and name.lower() in agg_text.lower():
                    patient_info['id'] = pid
                    patient_info['name'] = name
                    patient_info['code'] = code
                    if addr:
                        patient_info['address'] = addr
                    if age:
                        patient_info['age'] = age
                    break
        finally:
            try:
                conn.close()
            except Exception:
                pass

    # try to detect doctor info by scanning known doktores
    doctor_info = {"name": None, "address": None, "specialty": None}
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cur = conn.cursor()
        cur.execute("SELECT name, specialty, email FROM doktores")
        for name, specialty, email in cur.fetchall():
            if name and name.lower() in agg_text.lower():
                doctor_info['name'] = name
                doctor_info['specialty'] = specialty
                # address not available in doktores table; try to capture from text
                addr_m = re.search(re.escape(name) + r"[\w\W]{0,120}?([A-ZÄÖÜ][^\n\r]+\d{1,4}[^\n\r]*)", agg_text, re.I)
                if addr_m:
                    doctor_info['address'] = addr_m.group(1).strip()
                break
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # ensure template header contains required fields that user requested
    required_fields = [
        'Code', 'Region', 'Stützpunkt', 'Kunde', 'Patientenangaben', 'Adresse', 'Alter', 'Name', 'Geburtsdatum',
        'Behandelnder Arzt', 'Arzt Adresse', 'Arzt Spezialität', 'Anzahl', 'Medikament', 'Erste/Letzte Verabreichung',
        'Morgens', 'Mittags', 'Abends', 'Täglich'
    ]
    # append missing required fields to header_row (preserving existing order)
    for f in required_fields:
        if f not in header_row:
            header_row.append(f)

    # call AI extractor using headers and patient_info/doctor_info
    ai_csv = _ai_extract_prescriptions(agg_text, header_row, patient_info, doctor_info, required_fields)
    if ai_csv is None:
        return {"ok": False, "error": "No prescriptions found", "extracted_text": agg_text, "patient": patient_info, "doctor": doctor_info}

    # parse ai_csv and write into workbook (starting at row 2)
    rows = list(csv.reader([l for l in ai_csv.splitlines() if l.strip()]))
    # if first row is header that exactly matches template header, skip it
    ai_headers = [h.strip() for h in rows[0]] if rows else []
    if ai_headers and [h.strip() for h in ai_headers] == [h.strip() for h in header_row]:
        data_rows = rows[1:]
    else:
        # try to detect if first row is a header with overlapping columns - if so, map/reorder
        if ai_headers:
            # build mapping from desired header to ai header index (case-insensitive match)
            mapping = {}
            ai_lower = [h.lower() for h in ai_headers]
            for idx_h, desired in enumerate(header_row):
                try:
                    i = ai_lower.index(desired.strip().lower())
                    mapping[idx_h] = i
                except ValueError:
                    # try fuzzy: check if any ai header contains desired or vice versa
                    found = None
                    for j, ah in enumerate(ai_lower):
                        if desired.strip().lower() in ah or ah in desired.strip().lower():
                            found = j
                            break
                    if found is not None:
                        mapping[idx_h] = found
            if mapping:
                # transform data rows (skip ai header)
                transformed = []
                for r in rows[1:]:
                    newrow = [''] * len(header_row)
                    for dest_idx, src_idx in mapping.items():
                        if src_idx < len(r):
                            newrow[dest_idx] = r[src_idx]
                    transformed.append(newrow)
                data_rows = transformed
            else:
                data_rows = rows
        else:
            data_rows = rows

    # if patient_info present, fill patient-related columns when empty
    # detect indices for common patient columns (case-insensitive)
    header_lower = [h.strip().lower() for h in header_row]
    pid_idx = None
    pname_idx = None
    pcode_idx = None
    addr_idx = None
    age_idx = None
    dob_idx = None
    for i, h in enumerate(header_lower):
        if 'patient' in h and 'id' in h:
            pid_idx = i
        if 'name' in h and ('patient' in h or 'kunde' in h):
            pname_idx = i
        if 'code' in h or 'patient code' in h:
            pcode_idx = i
        if 'adres' in h or 'adresse' in h:
            addr_idx = i
        if 'alter' in h or 'age' in h:
            age_idx = i
        if 'geb' in h or 'birth' in h or 'dob' in h:
            dob_idx = i
    filled_rows = []
    for row in data_rows:
        cols = list(row)
        # extend row to header length if needed
        if len(cols) < len(header_row):
            cols += [''] * (len(header_row) - len(cols))
        if patient_info.get('id') and pid_idx is not None and (not cols[pid_idx] or cols[pid_idx].strip()==''):
            cols[pid_idx] = str(patient_info['id'])
        if patient_info.get('name') and pname_idx is not None and (not cols[pname_idx] or cols[pname_idx].strip()==''):
            cols[pname_idx] = patient_info['name']
        if patient_info.get('code') and pcode_idx is not None and (not cols[pcode_idx] or cols[pcode_idx].strip()==''):
            cols[pcode_idx] = str(patient_info['code'])
        if patient_info.get('address') and addr_idx is not None and (not cols[addr_idx] or cols[addr_idx].strip()==''):
            cols[addr_idx] = patient_info['address']
        if patient_info.get('age') and age_idx is not None and (not cols[age_idx] or cols[age_idx].strip()==''):
            cols[age_idx] = str(patient_info['age'])
        if patient_info.get('dob') and dob_idx is not None and (not cols[dob_idx] or cols[dob_idx].strip()==''):
            cols[dob_idx] = str(patient_info['dob'])
        # fill doctor info
        if doctor_info.get('name') and 'behandelnder arzt' in header_lower and (not cols[header_lower.index('behandelnder arzt')] or cols[header_lower.index('behandelnder arzt')].strip()==''):
            cols[header_lower.index('behandelnder arzt')] = doctor_info.get('name')
        if doctor_info.get('specialty') and 'arzt spezial' in ' '.join(header_lower) and any('spez' in h for h in header_lower):
            # attempt to find the specialty column
            for idx_h, hh in enumerate(header_lower):
                if 'spez' in hh or 'special' in hh:
                    if not cols[idx_h] or cols[idx_h].strip()=='':
                        cols[idx_h] = doctor_info.get('specialty')
                        break
        filled_rows.append(cols)
    data_rows = filled_rows

    # normalize boolean/frequency columns
    # detect indices
    tgl_idx = None
    morg_idx = None
    mitt_idx = None
    abend_idx = None
    anzahl_idx = None
    for i, h in enumerate(header_lower):
        if 'täg' in h or 'täglich' in h:
            tgl_idx = i
        if 'morg' in h:
            morg_idx = i
        if 'mitt' in h:
            mitt_idx = i
        if 'abend' in h:
            abend_idx = i
        if 'anz' in h:
            anzahl_idx = i

    def norm_bool(v: str) -> str:
        if v is None:
            return ''
        s = str(v).strip().lower()
        if s in ('ja', 'j', 'yes', 'y', '1', 'true', 'x', 'x'):
            return 'Ja'
        if s in ('nein', 'n', 'no', '0', 'false'):
            return 'Nein'
        return v

    def norm_count(v: str) -> str:
        if v is None:
            return ''
        s = str(v).strip().lower()
        if s == '':
            return ''
        # try to extract integer
        m = re.search(r"(\d+)", s)
        if m:
            return m.group(1)
        # common words
        if s in ('ja', 'j', 'x'):
            return '1'
        if s in ('nein', 'n'):
            return '0'
        return s

    # apply normalization
    normalized_rows = []
    for row in data_rows:
        newr = list(row)
        if tgl_idx is not None and tgl_idx < len(newr):
            newr[tgl_idx] = norm_bool(newr[tgl_idx])
        if morg_idx is not None and morg_idx < len(newr):
            newr[morg_idx] = norm_count(newr[morg_idx])
        if mitt_idx is not None and mitt_idx < len(newr):
            newr[mitt_idx] = norm_count(newr[mitt_idx])
        if abend_idx is not None and abend_idx < len(newr):
            newr[abend_idx] = norm_count(newr[abend_idx])
        if anzahl_idx is not None and anzahl_idx < len(newr):
            newr[anzahl_idx] = norm_count(newr[anzahl_idx])
        normalized_rows.append(newr)
    data_rows = normalized_rows

    # clear existing rows below header
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row)
    # write rows
    ridx = 2
    for row in data_rows:
        for cidx, val in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=val)
        ridx += 1

    out_buf = io.BytesIO()
    fname_out = f"{tpl_path.stem}_filled_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(out_buf)
    out_buf.seek(0)

    if as_zip:
        z = io.BytesIO()
        with zipfile.ZipFile(z, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(fname_out, out_buf.getvalue())
            if original_csv_bytes:
                zf.writestr(original_csv_name or 'original.csv', original_csv_bytes)
        z.seek(0)
        return StreamingResponse(z, media_type='application/zip', headers={"Content-Disposition": f"attachment; filename={tpl_path.stem}_filled.zip"})

    return StreamingResponse(io.BytesIO(out_buf.getvalue()), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={fname_out}"})


@app.get("/emails/extract_ui")
async def extract_ui():
    """Simple HTML form to browse and upload a file to /emails/extract_and_fill_template."""
    html = '''
    <html>
      <head><title>Extract & Fill</title></head>
      <body>
        <h1>Upload file to extract prescriptions</h1>
        <form action="/emails/extract_and_fill_template" method="post" enctype="multipart/form-data">
          <label>File: <input type="file" name="file"/></label><br/><br/>
          <label>Or paste text:<br/><textarea name="text" rows="10" cols="80"></textarea></label><br/><br/>
          <label>Template path (optional): <input name="template" size="60"/></label><br/>
          <label>Return ZIP: <input type="checkbox" name="as_zip" checked/></label><br/><br/>
          <input type="submit" value="Upload & Extract" />
        </form>
      </body>
    </html>
    '''
    return HTMLResponse(content=html)


@app.post("/examples/create_patient_samples")
async def create_patient_samples(count: int = 5):
    """Create `count` sample files (PDF + CSV + TXT) based on random patients and store them in `examples/samples/`."""
    samples_dir = Path(__file__).parent / 'examples' / 'samples'
    samples_dir.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, name, code FROM patients")
        patients = cur.fetchall()
        if not patients:
            raise HTTPException(status_code=400, detail="No patients available to generate samples")
        created = []
        for i in range(count):
            pid, name, code = random.choice(patients)
            base = f"sample_{pid}_{uuid.uuid4().hex[:6]}"

            # pick a random doctor to appear on the sample (if any)
            cur.execute("SELECT name FROM doktores")
            docs = [d[0] for d in cur.fetchall()]
            doc_name = random.choice(docs) if docs else "Dr. Example"

            # CSV (med list) - include Doctor column and a Notes column with extra filler
            csv_name = f"{base}.csv"
            csv_path = samples_dir / csv_name
            rows = [
                ["Patientenname", "Patienten-ID", "Medikament", "Dosierung", "Anzahl", "Arzt", "Anmerkungen"],
                [name, pid, "Aspirin", "100mg", "30", doc_name, "Beispielrezept. Ignorieren Sie Extratext."],
                [name, pid, "Paracetamol", "500mg", "20", doc_name, f"Nur für Testzwecke. RNG:{uuid.uuid4().hex[:6]}"]
            ]
            with csv_path.open('w', newline='', encoding='utf-8') as cf:
                w = csv.writer(cf)
                for r in rows:
                    w.writerow(r)

            # TXT - include doctor and some extra irrelevant text
            txt_name = f"{base}.txt"
            txt_path = samples_dir / txt_name
            txt_content = (
                f"Rezept für {name} (ID: {pid})\n"
                f"Verschrieben von: {doc_name}\n"
                "- Aspirin 100mg 30\n"
                "- Paracetamol 500mg 20\n"
                "\nHinweis: Dies sind Beispieldaten für Testzwecke. Bitte ignorieren Sie zusätzliche Zeilen.\n"
                f"Referenz: {uuid.uuid4().hex[:8]}\n"
                "---\n"
                "Zusätzlicher Text: Dies ist ein Beispiel-Dokument.\n"
            )
            txt_path.write_text(txt_content, encoding='utf-8')

            # PDF - add doctor and filler lines
            pdf_name = f"{base}.pdf"
            pdf_path = samples_dir / pdf_name
            if canvas:
                c = canvas.Canvas(str(pdf_path))
                c.drawString(72, 750, f"Rezept für {name} (ID: {pid})")
                c.drawString(72, 732, f"Verschrieben von: {doc_name}")
                c.drawString(72, 712, "- Aspirin 100mg 30")
                c.drawString(72, 698, "- Paracetamol 500mg 20")
                c.drawString(72, 660, "Hinweis: Nur für Testzwecke. Ignorieren Sie zusätzliche Zeilen.")
                c.drawString(72, 640, f"Ref: {uuid.uuid4().hex[:8]}")
                c.save()
            else:
                # fallback: write a simple text in PDF-like file
                pdf_path.write_text(txt_content, encoding='utf-8')

            created.append({"csv": str(csv_path), "txt": str(txt_path), "pdf": str(pdf_path), "patient_id": pid, "patient_name": name})
        return {"ok": True, "created": created}
    finally:
        conn.close()


@app.post("/emails/sort_unsorted")
async def sort_unsorted_emails(dry_run: bool = False, limit: int | None = None):
    """Auto-sort .eml files in `storage/emails/unsorted`.

    - Tries to detect a patient id (from Subject or body) or match patient name
      (from `patients` table) in the subject/body.
    - If a match is found, moves the .eml into `storage/emails/{patient_id}/`.
    - Parameters:
      - dry_run: if true, don't move files, only report what would be done
      - limit: optional max number of files to process
    """
    unsorted_dir = STORAGE_EMAILS_DIR / "unsorted"
    if not unsorted_dir.exists():
        return {"ok": True, "processed": 0, "moved": 0, "details": []}

    from email.parser import BytesParser
    parser = BytesParser(policy=policy.default)

    processed = 0
    moved = 0
    details = []

    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, name, code FROM patients")
        patients = cur.fetchall()  # (id, name, code)
        lower_names = [(pid, name, (name or '').lower(), (code or '').upper()) for pid, name, code in patients]

        for p in sorted(unsorted_dir.iterdir()):
            if limit is not None and processed >= limit:
                break
            if not p.is_file() or not p.name.lower().endswith('.eml'):
                continue
            processed += 1
            try:
                raw = p.read_bytes()
                msg = parser.parsebytes(raw)
                subject = (msg.get('Subject') or "")
                # assemble plain-text body
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type().startswith('text/'):
                            try:
                                body += part.get_content() or ""
                            except Exception:
                                continue
                else:
                    try:
                        body = msg.get_content() or ""
                    except Exception:
                        body = ""

                        # try to find numeric id
                id_m = re.search(r"\b(?:patient|id)\s*[:#-]?\s*(\d+)\b", subject, re.I)
                if not id_m:
                    id_m = re.search(r"\b(?:patient|id)\s*[:#-]?\s*(\d+)\b", body, re.I)
                patient_id = int(id_m.group(1)) if id_m else None
                method = None

                # also try to find patient code (5 digits) in subject/body
                code_m = re.search(r"\b(\d{5})\b", subject or "") or re.search(r"\b(\d{5})\b", body or "")
                patient_code = code_m.group(1) if code_m else None

                if not patient_id and not patient_code:
                    # try name match
                    combined = (subject + " " + body).lower()
                    for pid, name, lname in lower_names:
                        if lname in combined:
                            patient_id = pid
                            method = 'name'
                            break
                    if patient_id:
                        method = method or 'name'
                elif patient_id:
                    method = 'id'
                elif patient_code:
                    # resolve code to id using pre-fetched codes
                    for pid, name, lname, pcode in lower_names:
                        if pcode and pcode.upper() == patient_code.upper():
                            patient_id = pid
                            method = 'code'
                            break

                if patient_id:
                    # use patient code as folder when available
                    cur.execute("SELECT code FROM patients WHERE id = ?", (patient_id,))
                    code_row = cur.fetchone()
                    folder = code_row[0] if code_row and code_row[0] else str(patient_id)
                    dest_dir = STORAGE_EMAILS_DIR / folder
                    dest_dir.mkdir(parents=True, exist_ok=True)
                    dest_path = dest_dir / p.name
                    base, ext = os.path.splitext(p.name)
                    counter = 1
                    while dest_path.exists():
                        dest_path = dest_dir / f"{base}_{counter}{ext}"
                        counter += 1
                    if not dry_run:
                        shutil.move(str(p), str(dest_path))
                    details.append({"file": p.name, "from": str(p), "to": str(dest_path), "patient_id": patient_id, "method": method, "moved": not dry_run})
                    if not dry_run:
                        moved += 1
                else:
                    details.append({"file": p.name, "from": str(p), "to": None, "patient_id": None, "method": None, "moved": False})
            except Exception as e:
                details.append({"file": p.name, "error": str(e)})

        return {"ok": True, "processed": processed, "moved": moved, "details": details}
    finally:
        conn.close()


@app.get("/patients/{patient_identifier}/emails")
async def list_patient_emails(patient_identifier: str):
    """List .eml files for a patient identifier (code or numeric id)."""
    try:
        pid, pcode, pname = _resolve_patient(patient_identifier)
        patient_dir = STORAGE_EMAILS_DIR / (pcode or str(pid))
    except HTTPException:
        # unknown patient
        raise HTTPException(status_code=404, detail="Patient not found")

    if not patient_dir.exists():
        return {"emails": []}

    emails = []
    from email.parser import BytesParser
    parser = BytesParser(policy=policy.default)

    for p in sorted(patient_dir.iterdir()):
        if not p.is_file() or not p.name.lower().endswith('.eml'):
            continue
        try:
            with p.open('rb') as f:
                msg = parser.parse(f)
            subject = msg.get('Subject')
            _from = msg.get('From')
            _to = msg.get('To')
            date = msg.get('Date')
        except Exception:
            subject = None
            _from = None
            _to = None
            date = None
        emails.append({
            "filename": p.name,
            "subject": subject,
            "from": _from,
            "to": _to,
            "date": date,
            "size": p.stat().st_size,
            "modified": p.stat().st_mtime
        })
    return {"patient": {"id": pid, "code": pcode, "name": pname}, "emails": emails}


@app.get("/patients/{patient_identifier}/emails/{filename}/download")
async def download_eml_file(patient_identifier: str, filename: str):
    """Download a stored .eml by patient identifier and filename."""
    try:
        pid, pcode, pname = _resolve_patient(patient_identifier)
        patient_dir = STORAGE_EMAILS_DIR / (pcode or str(pid))
    except HTTPException:
        raise HTTPException(status_code=404, detail="Patient not found")
    path = patient_dir / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Email file not found")
    return FileResponse(path, media_type="message/rfc822", filename=path.name)


@app.get("/patients/{patient_identifier}/emails/{filename}/attachment")
async def download_email_attachment_by_file(patient_identifier: str, filename: str):
    """Download the first attachment from a stored .eml file for a patient."""
    try:
        pid, pcode, pname = _resolve_patient(patient_identifier)
        patient_dir = STORAGE_EMAILS_DIR / (pcode or str(pid))
    except HTTPException:
        raise HTTPException(status_code=404, detail="Patient not found")
    eml_path = patient_dir / filename
    if not eml_path.exists() or not eml_path.is_file():
        raise HTTPException(status_code=404, detail="Email file not found")
    from email.parser import BytesParser
    parser = BytesParser(policy=policy.default)
    with eml_path.open('rb') as f:
        msg = parser.parse(f)
    for part in msg.iter_attachments():
        fname = part.get_filename()
        payload = part.get_content()
        if isinstance(payload, bytes):
            return StreamingResponse(io.BytesIO(payload), media_type=part.get_content_type(), headers={"Content-Disposition": f'attachment; filename="{fname}"'})
        else:
            data = payload.encode('utf-8')
            return StreamingResponse(io.BytesIO(data), media_type=part.get_content_type(), headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    raise HTTPException(status_code=404, detail="No attachment found in the .eml file")


#-- AI CSV endpoint ---
@app.post("/csv")
async def csv(message: str | None = None, task: str | None = None, csv: str | None = None):
    """General chat endpoint. If `task` == 'parse_medlist' (or 'parse_csv'), the endpoint
    expects a `csv` body and returns both the original CSV and an AI-standardized CSV
    that extracts medication rows. Otherwise, behaves like a normal chat.
    """
    if task and task.lower() in ("parse_medlist", "parse_csv"):
        if not csv:
            raise HTTPException(status_code=400, detail="Missing 'csv' payload for parse_medlist task")
        original = csv
        # try simple normalization
        normalized = original.replace('\r\n', '\n').replace('\r', '\n')
        ai_csv = _ai_parse_csv_for_medlist(normalized)
        if ai_csv is None:
            return {"ok": False, "error": "AI did not return a medlist extract", "original_csv": original}
        return {"ok": True, "original_csv": original, "standardized_csv": ai_csv}

    # fallback to plain chat behavior
    if not message:
        raise HTTPException(status_code=400, detail="Provide 'message' for chat or use task='parse_medlist' with 'csv' parameter")
    response = client.chat.completions.create(
        model="openai/gpt-oss-20b",
        messages=[
            {"role": "system", "content": "answer in slang."},
            {"role": "user", "content": message}
        ]
    )
    return {"response": response.choices[0].message.content}

@app.patch("/messages_db/{message_id}")
async def update_message_db(message_id: int, data: dict = Body(...)):
    """Update a chat message's status or other fields."""
    if not DB_PATH.exists():
        raise HTTPException(status_code=500, detail="Database not initialized. POST /init_db first.")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    try:
        # Build dynamic update query
        fields = []
        values = []
        for k, v in data.items():
            if k in ['status', 'is_important', 'content', 'due_date']:
                fields.append(f"{k} = ?")
                values.append(v)
        
        if not fields:
            raise HTTPException(status_code=400, detail="No valid fields to update")
            
        values.append(message_id)
        query = f"UPDATE chat_messages SET {', '.join(fields)} WHERE id = ?"
        cur.execute(query, values)
        conn.commit()
        
        return {"ok": True}
    finally:
        conn.close()

@app.post("/Chat")
async def chat_endpoint(message: str):
    """Simple chat endpoint that forwards the message to the AI model and returns the response."""
    if not message:
        raise HTTPException(status_code=400, detail="Provide 'message' for chat")
    response = client.chat.completions.create(
        model="openai/gpt-oss-20b",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": message}
        ]
    )
    return {"response": response.choices[0].message.content}